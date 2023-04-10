from openpyxl import load_workbook

workbook = load_workbook("example.xlsx")

sheet = workbook.active

celula = sheet.cell(row=1, column=1)
celula.value = 50

print(celula.value)

