from openpyxl import Workbook

# Criar um novo arquivo de Excel
wb = Workbook()

# Selecionar a planilha ativa
ws = wb.active

# Definir a largura da coluna A para 25
ws.column_dimensions['A'].width = 25
ws.row_dimensions[1].height = 25

# Adicionar alguns dados à planilha
ws['A1'] = 'Nome'
ws['B1'] = 'Idade'
ws['C1'] = 'Cidade'

ws['A2'] = 'João'
ws['B2'] = 30
ws['C2'] = 'São Paulo'

ws['A3'] = 'Maria'
ws['B3'] = 25
ws['C3'] = 'Rio de Janeiro'

# Salvar a planilha em um arquivo Excel
wb.save('exemplo_planilha.xlsx')