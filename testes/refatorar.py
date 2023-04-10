from openpyxl import load_workbook
from datetime import datetime
import os
import shutil
from math import ceil


print(__name__)


def metadados(mestra, base, coluna, data):
    # Ler planilha base para verificar largura das celulas na coluna de preenchimento:
    wb_base = load_workbook(base)
    sheet_base = wb_base.sheetnames[0]
    sheet_b = wb_base[sheet_base]

    # Para converter as letras das colunas em maiusculas
    coluna = coluna.upper()

    # Variavel para pegar o valor da largura das colunas
    len_col = sheet_b.column_dimensions[coluna].width

    # Numero máximo de caracteres aproximado pela largura da colula
    n_carac = int(len_col)
    len_row_1 = 30  # Padrão excel

    # Ler planilha com dados para atualização:
    wb = load_workbook(mestra)

    # Ler nome da aba da planilha e ativar edição:
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    ws = wb.active

    # Numero de linhas e colunas maximas da planilha mestra:
    mr = ws.max_row
    mc = ws.max_column

    # Criar listas com caminhos e nomes dos metadados a serem gerados:
    paths_list = []
    lista_metadados = []
    # Fixaria coluna 3 (colunas correspondentes Ã s planilhas)

    for j in range(3, mc + 1):
        # Iniciando da linha 6 para caminho #Fixaria na 1
        path = ws.cell(row=1, column=j)
        paths_list.append(path.value)

        # Iniciando da linha 7 para desconsiderar explicações #Fixar na 2
        c = ws.cell(row=2, column=j)
        lista_metadados.append(c.value)

    # Criar planilha de metadados a ser atualizada a partir da base:

    for nome_metadado in lista_metadados:
        # Gerar caminho e cópia da atualização:
        idx = lista_metadados.index(nome_metadado)
        path = paths_list[idx]

        path_metadado_prov = os.path.join(path, "metadado.xlsx")

        # Deixar data como input

        path_metadado = os.path.join(
            path, "metadados_" + nome_metadado + "_hidrobr_" + data + ".xlsx"
        )

        try:
            shutil.copy(base, path_metadado_prov)
            print("File copied successfully.")
        # If source and destination are same
        except shutil.SameFileError:
            print("Source and destination represents the same file.")
        # If there is any permission issue
        except PermissionError:
            print("Permission denied.")
        # For other errors
        except:
            print("Error occurred while copying file.")

        # Abrir arquivo gerado:
        try:
            wb_m = load_workbook(path_metadado_prov)
        except:
            continue

        sheet_m = wb_m.sheetnames[0]
        ws_m = wb_m[sheet_m]
        ws_m = wb_m.active

        # Renomear aba da planilha de acordo com nome do arquivo de metadado:
        ws_m.title = nome_metadado

        # Atualizar campos da planilha:

        # 1)De acordo com dados gerais (coluna 3):
        for i in range(3, mr + 1):
            c = ws.cell(row=i, column=2)
            # Compatibilizar inÃ­cio na planilha de metadados #cell como input (inÃ­cio do preenchimento)
            cell = i + 3

            val = c.value

            if type(val) == datetime:
                # val = c.value
                val = val.strftime("%d/%m/%Y")
                ws_m[coluna + str(cell)] = val
            else:
                # Nova versão
                # Verificar necessidade de ajuste da celula:
                if type(val) == str and len(val) > n_carac:
                    n_rows = ceil(len(val) / n_carac)
                    len_row = ws_m.row_dimensions[cell].height
                    len_cell = n_rows * len_row_1
                    if len_cell > len_row:
                        ws_m.row_dimensions[cell].height = len_cell
                    else:
                        pass
                else:
                    pass

                ws_m[coluna + str(cell)] = val

        # 2)De acordo com coluna especi­fica do metadado:
        col = lista_metadados.index(nome_metadado) + 3  # Fixar para +3
        # ws_m['C2'] = ws.cell(row = 2, column = col).value #TÃ­tulo
        for i in range(3, mr + 1):
            c = ws.cell(row=i, column=col)
            cell = i + 3  # Compatibilizar inÃ­cio na planilha de metadados

            val = c.value

            if val != None:
                if type(val) == datetime:
                    # val = c.value
                    val = val.strftime("%d/%m/%Y")
                    ws_m[coluna + str(cell)] = val
                else:
                    # Verificar necessidade de ajuste da cÃ©lula:
                    if type(val) == str and len(val) > n_carac:
                        n_rows = ceil(len(val) / n_carac)
                        len_row = ws_m.row_dimensions[cell].height
                        len_cell = n_rows * len_row_1
                        if len_cell > len_row:
                            ws_m.row_dimensions[cell].height = len_cell
                        else:
                            pass
                    else:
                        pass

                    ws_m[coluna + str(cell)] = val

        # Excluir as planilhas

        if os.path.exists("metadados_" + nome_metadado + "_hidrobr_" + data + ".xlsx"):
            os.remove("metadados_" + nome_metadado + "_hidrobr_" + data + ".xlsx")

        # Salvar/Criar nova planilha
        wb_m.save(path_metadado_prov)
        try:
            os.rename(path_metadado_prov, path_metadado)
        except Exception as e:
            print(e, "\n")
